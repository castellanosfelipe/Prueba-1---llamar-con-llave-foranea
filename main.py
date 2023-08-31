import tkinter as tk
from tkinter import ttk, messagebox, font
import mysql.connector
from tkcalendar import DateEntry
from openpyxl import Workbook
import os

class AppRegistroHoras:
    def __init__(self, root):
        self.root = root
        self.root.title("Registro de horas")

        self.connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123456789",
            database="crudd",
            port='3306'
        )
        self.cursor = self.connection.cursor()

        bold_font = font.Font(weight="bold")
        
        self.totales_label = tk.Label(self.root, text="Totales de horas por proyecto:", font=bold_font)
        self.totales_label.grid(row=9, column=0, columnspan=2, padx=5, pady=5)


        self.totales_treeview = ttk.Treeview(self.root, columns=("Proyecto", "Total Horas"), show="headings", height=3)
        self.totales_treeview.heading("Proyecto", text="Proyecto")
        self.totales_treeview.heading("Total Horas", text="Total Horas")
        self.totales_treeview.grid(row=10, column=0, columnspan=2, padx=5, pady=3)


        self.diseño_interfaz()
        self.cargar_proyectos()
        self.cargar_empleados()
        self.crear_tabla()
        self.totales_proyecto = {}  # Diccionario para almacenar los totales de horas por proyecto
        
    def calcular_totales_proyecto(self):
        self.totales_proyecto = {}  # Reiniciar el diccionario de totales
        for registro in self.tabla.get_children():
            proyecto = self.tabla.item(registro)["values"][2]  # Obtener el nombre del proyecto
            horas = float(self.tabla.item(registro)["values"][4])  # Obtener las horas del registro
            if proyecto in self.totales_proyecto:
                self.totales_proyecto[proyecto] += horas
            else:
                self.totales_proyecto[proyecto] = horas

        # Actualizar los valores en el Treeview de totales
        self.totales_treeview.delete(*self.totales_treeview.get_children())
        for proyecto, total_horas in self.totales_proyecto.items():
            self.totales_treeview.insert("", "end", values=[proyecto, total_horas])

      
    def validar_formato_fecha(self, fecha):
        try:
            if len(fecha.split("-")) == 3:
                y, m, d = map(int, fecha.split("-"))
                if 1000 <= y <= 9999 and 1 <= m <= 12 and 1 <= d <= 31:
                    return True
            return False
        except ValueError:
            return False
    
    def abrir_excel(self, archivo_excel):
        try:
            os.system(f'start excel "{archivo_excel}"')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")

    def exportar_a_excel(self):
            wb = Workbook()
            ws = wb.active

            # Agregar encabezados de columnas
            columnas = ["ID", "Empleado", "Proyecto", "Requerimiento", "No. Horas", "Fecha", "Descripción"]
            for col_num, columna in enumerate(columnas, start=1):
                ws.cell(row=1, column=col_num, value=columna)

            # Agregar datos de la tabla
            for row_num, registro in enumerate(self.tabla.get_children(), start=2):
                values = self.tabla.item(registro)["values"]
                for col_num, value in enumerate(values, start=1):
                    ws.cell(row=row_num, column=col_num, value=value)

            # Guardar el archivo Excel
            archivo_excel = "registros_horas.xlsx"
            wb.save(archivo_excel)
            messagebox.showinfo("Éxito", f"Datos exportados a {archivo_excel}")
            self.abrir_excel(archivo_excel)

    def diseño_interfaz(self):
        self.empleado_label = tk.Label(self.root, text="Ing. Colaborador")
        self.empleado_label.grid(row=0, column=0, padx=10, pady=5)

        self.empleado_combobox = ttk.Combobox(self.root)
        self.empleado_combobox.grid(row=0, column=1, padx=10, pady=5)

        self.proyecto_label = tk.Label(self.root, text="Proyecto")
        self.proyecto_label.grid(row=1, column=0, padx=10, pady=5)

        self.proyecto_combobox = ttk.Combobox(self.root)
        self.proyecto_combobox.grid(row=1, column=1, padx=10, pady=5)
        self.proyecto_combobox.bind("<<ComboboxSelected>>", self.cargar_requerimientos)

        self.requerimiento_label = tk.Label(self.root, text="Requerimiento")
        self.requerimiento_label.grid(row=2, column=0, padx=5, pady=5)

        self.requerimiento_combobox = ttk.Combobox(self.root)
        self.requerimiento_combobox.grid(row=2, column=1, padx=10, pady=5)

        self.empleado_combobox.bind("<<ComboboxSelected>>", self.actualizar_tabla_por_empleado)
        self.proyecto_combobox.bind("<<ComboboxSelected>>", self.cargar_requerimientos)

        self.hora_label = tk.Label(self.root, text="No. Horas")
        self.hora_label.grid(row=3, column=0, padx=10, pady=5)
        self.hora_entry = tk.Entry(self.root)
        self.hora_entry.grid(row=3, column=1, padx=10, pady=5)
        self.hora_entry.bind("<KeyRelease>", self.validar_hora_numeros)  # Agregar validación

        self.fecha_label = tk.Label(self.root, text="Fecha")
        self.fecha_label.grid(row=4, column=0, padx=10, pady=5)
        self.fecha_entry = DateEntry(self.root, date_pattern="YYYY-MM-DD")
        self.fecha_entry.grid(row=4, column=1, padx=10, pady=5)

        self.descripcion_label = tk.Label(self.root, text="Descripción")
        self.descripcion_label.grid(row=5, column=0, padx=10, pady=5)
        self.descripcion_entry = tk.Entry(self.root)
        self.descripcion_entry.grid(row=5, column=1, padx=10, pady=5)

        self.registrar_button = tk.Button(self.root, text="Registrar", command=self.registrar_registro, bg="green")
        self.registrar_button.grid(row=6, column=1, padx=50, pady=5,columnspan=1,  sticky="we")
        
        self.modificar_button = tk.Button(self.root, text="Modificar", command=self.modificar_registro, bg="yellow")
        self.modificar_button.grid(row=6, column=0, padx=10, pady=5, columnspan=1, sticky="we")
        
        self.exportar_button = tk.Button(self.root, text="Exportar a excel", command=self.exportar_a_excel)
        self.exportar_button.grid(row=7, column=0, padx=10, pady=5, sticky="we", columnspan=2)

        self.tabla = ttk.Treeview(self.root, columns=("ID", "Empleado", "Proyecto", "Requerimiento", "Hora (HH:MM:SS)", "Fecha (YYYY-MM-DD)", "Descripción"), show="headings", height=5)
        self.tabla.heading("ID", text="ID")
        self.tabla.heading("Empleado", text="Empleado")
        self.tabla.heading("Proyecto", text="Proyecto")
        self.tabla.heading("Requerimiento", text="Requerimiento")
        self.tabla.heading("Hora (HH:MM:SS)", text="No. Horas")
        self.tabla.heading("Fecha (YYYY-MM-DD)", text="Fecha")
        self.tabla.heading("Descripción", text="Descripción")
        self.tabla.grid(row=8, column=0, columnspan=2, padx=5, pady=10)
        
        self.tabla.column("ID", width=30)  # Ajustar el ancho de la columna "ID"
        self.tabla.column("Empleado", width=100)  # Ajustar el ancho de la columna "Empleado"
        self.tabla.column("Proyecto", width=100)  # Ajustar el ancho de la columna "Proyecto"
        self.tabla.column("Requerimiento", width=150)  # Ajustar el ancho de la columna "Requerimiento"
        self.tabla.column("Hora (HH:MM:SS)", width=80) 
        self.tabla.column("Fecha (YYYY-MM-DD)", width=80)
        self.tabla.column("Descripción", width=200) 
        
    def validar_hora_numeros(self, event):
        entrada = self.hora_entry.get()

        try:
            if entrada:
                float(entrada)  # Intenta convertir la entrada a un número decimal
        except ValueError:
            self.hora_entry.delete(len(entrada) - 1, tk.END)
      
    def cargar_empleados(self):
        query = "SELECT id, nombre_empleado FROM empleados"
        self.cursor.execute(query)
        empleados = self.cursor.fetchall()

        self.empleado_combobox["values"] = [empleado[1] for empleado in empleados]

    def cargar_proyectos(self):
        query = "SELECT id, nombre_proyecto FROM proyectos"
        self.cursor.execute(query)
        proyectos = self.cursor.fetchall()

        self.proyecto_combobox["values"] = [proyecto[1] for proyecto in proyectos]

    def cargar_requerimientos(self, event):
        proyecto_nombre = self.proyecto_combobox.get()
        query = "SELECT requerimientos.id, requerimientos.nombre_requerimiento FROM requerimientos JOIN proyectos ON requerimientos.proyecto_id = proyectos.id WHERE proyectos.nombre_proyecto = %s"
        self.cursor.execute(query, (proyecto_nombre,))
        requerimientos = self.cursor.fetchall()

        self.requerimiento_combobox["values"] = [requerimiento[1] for requerimiento in requerimientos]

    def crear_tabla(self):
        self.tabla.delete(*self.tabla.get_children())  # Limpiar la tabla
        # Obtener el nombre del empleado seleccionado en la lista desplegable
        empleado_nombre = self.empleado_combobox.get()

        query = "SELECT RegistroHoras.id, nombre_empleado, nombre_proyecto, nombre_requerimiento, hora, fecha, descripcion FROM RegistroHoras JOIN proyectos ON RegistroHoras.proyecto_id = proyectos.id JOIN requerimientos ON RegistroHoras.requerimiento_id = requerimientos.id JOIN empleados ON RegistroHoras.empleado_id = empleados.id"

        if empleado_nombre:  # Si se ha seleccionado un empleado, agregar el filtro a la consulta
            query += " WHERE nombre_empleado = %s"

        query += " ORDER BY RegistroHoras.id ASC"

        if empleado_nombre:  # Si se ha seleccionado un empleado, proporcionar el valor del filtro
            self.cursor.execute(query, (empleado_nombre,))
        else:
            self.cursor.execute(query)

        registros = self.cursor.fetchall()

        for registro in registros:
            registro = list(registro)
            registro[4] = float(registro[4])
            self.tabla.insert("", "end", values=registro)

        self.calcular_totales_proyecto()  # Agregar esta línea para calcular y mostrar los totales


    def actualizar_tabla_por_empleado(self, event):
            self.crear_tabla()
    
            
    def registrar_registro(self):
        empleado_nombre = self.empleado_combobox.get()
        proyecto_nombre = self.proyecto_combobox.get()
        requerimiento_nombre = self.requerimiento_combobox.get()
        hora_entrada = self.hora_entry.get()
        fecha = self.fecha_entry.get()
        descripcion = self.descripcion_entry.get()

        if not empleado_nombre or not proyecto_nombre or not requerimiento_nombre or not hora_entrada or not fecha or not descripcion or not self.validar_formato_fecha(self.fecha_entry.get()):
            messagebox.showinfo("Alerta", "Por favor complete todos los campos correctamente antes de registrar.")
            return

        try:
            hora = float(hora_entrada)
        except ValueError:
            messagebox.showerror("Error", "El valor de la hora no es válido.")
            return

        query_empleado_id = "SELECT id FROM empleados WHERE nombre_empleado = %s"
        self.cursor.execute(query_empleado_id, (empleado_nombre,))
        empleado_id = self.cursor.fetchone()[0]

        query_proyecto_id = "SELECT id FROM proyectos WHERE nombre_proyecto = %s"
        self.cursor.execute(query_proyecto_id, (proyecto_nombre,))
        proyecto_id = self.cursor.fetchone()[0]

        query_requerimiento_id = "SELECT id FROM requerimientos WHERE nombre_requerimiento = %s"
        self.cursor.execute(query_requerimiento_id, (requerimiento_nombre,))
        requerimiento_id = self.cursor.fetchone()[0]

        query = "INSERT INTO RegistroHoras (empleado_id, proyecto_id, requerimiento_id, hora, fecha, descripcion) VALUES (%s, %s, %s, %s, %s, %s)"
        values = (empleado_id, proyecto_id, requerimiento_id, hora, self.fecha_entry.get(), self.descripcion_entry.get())

        self.cursor.execute(query, values)
        self.connection.commit()

        self.empleado_combobox.set("")
        self.proyecto_combobox.set("")
        self.requerimiento_combobox.set("")
        self.hora_entry.delete(0, tk.END)
        self.fecha_entry.delete(0, tk.END)
        self.descripcion_entry.delete(0, tk.END)
        messagebox.showinfo("Éxito", "Registro de horas agregado con éxito.")
        self.crear_tabla()

        self.calcular_totales_proyecto()

    def modificar_registro(self):
        selected_item = self.tabla.selection()
        if not selected_item:
            messagebox.showinfo("Selección", "Por favor seleccione un registro a modificar.")
            return

        registro_id = self.tabla.item(selected_item)["values"][0]
        nuevo_empleado_nombre = self.empleado_combobox.get()
        nuevo_proyecto_nombre = self.proyecto_combobox.get()
        nuevo_requerimiento_nombre = self.requerimiento_combobox.get()
        nueva_hora = self.hora_entry.get()
        nueva_fecha = self.fecha_entry.get()
        nueva_descripcion = self.descripcion_entry.get()

        query_empleado_id = "SELECT id FROM empleados WHERE nombre_empleado = %s"
        self.cursor.execute(query_empleado_id, (nuevo_empleado_nombre,))
        nuevo_empleado_id = self.cursor.fetchone()[0]

        query_proyecto_id = "SELECT id FROM proyectos WHERE nombre_proyecto = %s"
        self.cursor.execute(query_proyecto_id, (nuevo_proyecto_nombre,))
        nuevo_proyecto_id = self.cursor.fetchone()[0]

        query_requerimiento_id = "SELECT id FROM requerimientos WHERE nombre_requerimiento = %s"
        self.cursor.execute(query_requerimiento_id, (nuevo_requerimiento_nombre,))
        nuevo_requerimiento_id = self.cursor.fetchone()[0]

        query = "UPDATE RegistroHoras SET empleado_id = %s, proyecto_id = %s, requerimiento_id = %s, hora = %s, fecha = %s, descripcion = %s WHERE id = %s"
        values = (nuevo_empleado_id, nuevo_proyecto_id, nuevo_requerimiento_id, nueva_hora, nueva_fecha, nueva_descripcion, registro_id)

        self.cursor.execute(query, values)
        self.connection.commit()
        
    # Limpiar los campos después de la modificación exitosa
        self.empleado_combobox.set("")
        self.proyecto_combobox.set("")
        self.requerimiento_combobox.set("")
        self.hora_entry.delete(0, tk.END)
        self.fecha_entry.delete(0, tk.END)
        self.descripcion_entry.delete(0, tk.END) 

        messagebox.showinfo("Éxito", "Registro de horas modificado con éxito.")
        self.crear_tabla()
        self.calcular_totales_proyecto()

if __name__ == "__main__":
    root = tk.Tk()
    app = AppRegistroHoras(root)
    root.geometry("755x540")
    root.resizable(False, False)
    root.mainloop()
